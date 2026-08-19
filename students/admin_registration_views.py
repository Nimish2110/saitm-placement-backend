from django.conf import settings
from django.contrib.auth import get_user_model
from django.core.mail import send_mail
from django.db import transaction
from django.utils import timezone
from rest_framework import generics, status
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_simplejwt.tokens import RefreshToken

from users.permissions import IsAdmin
from .models import StudentProfile
from .serializers import AdminStudentListSerializer, StudentInviteSerializer, CompleteRegistrationSerializer

User = get_user_model()

REQUIRED_HEADERS = ["Name", "Roll Number", "College Email", "Phone Number"]


def _tokens_for(user):
    refresh = RefreshToken.for_user(user)
    return {"access": str(refresh.access_token), "refresh": str(refresh)}


class AdminStudentListView(generics.ListAPIView):
    """GET /api/students/admin/list/ — every student, registered or still pending completion."""
    serializer_class = AdminStudentListSerializer
    permission_classes = [IsAuthenticated, IsAdmin]
    queryset = StudentProfile.objects.all().order_by("-created_at")


class BulkImportStudentsView(APIView):
    """
    POST /api/students/admin/bulk-import/
    Admin uploads an .xlsx with exactly 4 columns: Name, Roll Number,
    College Email, Phone Number (in that order, header row required).
    Creates an inactive User + a StudentProfile per row. Duplicates
    (by roll number or email) are skipped, not overwritten.
    """
    permission_classes = [IsAuthenticated, IsAdmin]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        uploaded_file = request.FILES.get("file")
        if not uploaded_file:
            return Response({"detail": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST)
        if not uploaded_file.name.endswith((".xlsx", ".xls")):
            return Response({"detail": "File must be an Excel (.xlsx) file."}, status=status.HTTP_400_BAD_REQUEST)

        from openpyxl import load_workbook

        try:
            wb = load_workbook(uploaded_file, read_only=True, data_only=True)
            ws = wb.active
        except Exception:
            return Response({"detail": "Could not read that file — make sure it's a valid .xlsx export."}, status=status.HTTP_400_BAD_REQUEST)

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return Response({"detail": "The file is empty."}, status=status.HTTP_400_BAD_REQUEST)

        header = [str(c).strip() if c else "" for c in rows[0]]
        if header[:4] != REQUIRED_HEADERS:
            return Response({
                "detail": f"The first row must be exactly these 4 column headers, in order: {', '.join(REQUIRED_HEADERS)}. Found: {', '.join(header[:4]) or '(empty)'}",
            }, status=status.HTTP_400_BAD_REQUEST)

        created, skipped, errors = 0, 0, []

        for i, row in enumerate(rows[1:], start=2):
            if not row or all(c is None for c in row[:4]):
                continue
            name, roll_no, email, phone = (str(c).strip() if c is not None else "" for c in row[:4])

            if not name or not roll_no or not email:
                errors.append(f"Row {i}: missing name, roll number, or email — skipped.")
                continue

            if StudentProfile.objects.filter(roll_no=roll_no).exists() or StudentProfile.objects.filter(college_email__iexact=email).exists():
                skipped += 1
                continue
            if User.objects.filter(email__iexact=email).exists():
                skipped += 1
                continue

            with transaction.atomic():
                user = User.objects.create(
                    username=email, email=email, first_name=name,
                    role=User.Role.STUDENT, is_active=False,
                )
                user.set_unusable_password()
                user.save()
                StudentProfile.objects.create(
                    user=user, full_name=name, roll_no=roll_no,
                    college_email=email, phone=phone or "",
                    home_address="", current_residence="",
                )
            created += 1

        return Response({
            "created": created, "skipped": skipped, "errors": errors,
            "detail": f"{created} students added, {skipped} already existed and were skipped.",
        }, status=status.HTTP_201_CREATED)


class SendInvitesView(APIView):
    """
    POST /api/students/admin/send-invites/
    Emails every student who hasn't completed registration yet a unique
    invite link to their college email. Sends directly within the request
    (not a background thread — those aren't reliable in a gunicorn/WSGI app
    and can silently die before finishing) using ONE shared SMTP connection
    for speed. Reports back exactly how many actually succeeded.
    """
    permission_classes = [IsAuthenticated, IsAdmin]

    def post(self, request):
        from django.core.mail import get_connection, EmailMessage

        pending = StudentProfile.objects.filter(registration_completed=False)
        sent = 0
        failed = 0
        connection = get_connection(fail_silently=False)
        connection.open()

        for profile in pending:
            link = f"{settings.FRONTEND_URL}/complete-registration/{profile.invite_token}"
            try:
                email = EmailMessage(
                    subject="Welcome to the SAITM Placement Portal — Complete Your Registration",
                    body=(
                        f"Hi {profile.full_name},\n\n"
                        f"You have successfully registered for the SAITM Placement Portal.\n\n"
                        f"Click the link below to set your password and complete your profile:\n{link}\n\n"
                        f"This link is unique to you — please don't share it.\n\n"
                        f"— SAITM T&P Cell"
                    ),
                    from_email=None,
                    to=[profile.college_email],
                    connection=connection,
                )
                email.send(fail_silently=False)
                profile.invite_sent_at = timezone.now()
                profile.save()
                sent += 1
            except Exception:
                failed += 1
                continue

        connection.close()

        detail = f"Sent to {sent} of {sent + failed} student(s)."
        if failed:
            detail += f" {failed} failed — check they have a valid college email on file."

        return Response({"sent": sent, "failed": failed, "detail": detail})
    
class InviteDetailView(APIView):
    """GET /api/students/invite/<token>/ — public. Shows the pre-filled data before the student fills in the rest."""
    permission_classes = [AllowAny]

    def get(self, request, token):
        try:
            profile = StudentProfile.objects.get(invite_token=token)
        except StudentProfile.DoesNotExist:
            return Response({"detail": "This invite link is invalid."}, status=status.HTTP_404_NOT_FOUND)
        if profile.registration_completed:
            return Response({"detail": "This registration is already complete. Please log in instead."}, status=status.HTTP_400_BAD_REQUEST)
        return Response(StudentInviteSerializer(profile).data)


class CompleteRegistrationView(APIView):
    """POST /api/students/invite/<token>/complete/ — public. Sets password + profile details, then logs them in."""
    permission_classes = [AllowAny]

    def post(self, request, token):
        try:
            profile = StudentProfile.objects.get(invite_token=token)
        except StudentProfile.DoesNotExist:
            return Response({"detail": "This invite link is invalid."}, status=status.HTTP_404_NOT_FOUND)
        if profile.registration_completed:
            return Response({"detail": "This registration is already complete. Please log in instead."}, status=status.HTTP_400_BAD_REQUEST)

        serializer = CompleteRegistrationSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        user = profile.user
        user.set_password(data["password"])
        user.is_active = True
        user.save()

        if data.get("phone"):
            profile.phone = data["phone"]
        profile.personal_email = data.get("personal_email", "")
        profile.home_address = data["home_address"]
        profile.current_residence = data["current_residence"]
        profile.course = data["course"]
        profile.batch = data["batch"]
        cgpa = data.get("cgpa")
        profile.cgpa = cgpa if cgpa else None
        profile.backlogs = data.get("backlogs", 0)
        profile.tenth_percentage = data["tenth_percentage"]
        profile.twelfth_percentage = data["twelfth_percentage"]
        profile.achievements = data.get("achievements", "")
        profile.certifications = data.get("certifications", "")
        profile.linkedin = data.get("linkedin", "")
        profile.github = data.get("github", "")
        profile.registration_completed = True
        profile.email_verified = True
        profile.save()

        return Response({"role": "student", **_tokens_for(user)}, status=status.HTTP_200_OK)