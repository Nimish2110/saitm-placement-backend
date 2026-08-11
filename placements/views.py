from rest_framework import generics
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework import status
from django.utils import timezone
from django.core.mail import send_mail

from users.permissions import IsStudent, IsPlacementManager
from students.models import StudentProfile, StudentDocument
from students.views import MAX_DOCS_BY_TYPE, DEFAULT_MAX_DOCS
from .models import Drive, Application, ResumeSampleTemplate
from .serializers import DriveSerializer, ApplicationSerializer, ApplicationFormSerializer, ResumeSampleTemplateSerializer


class DriveListCreateView(generics.ListCreateAPIView):
    """
    GET  -> students only see OPEN drives matching their own course AND batch.
            Placement Managers (or any non-student) see all open drives.
    POST -> Placement Manager only. Publishing a drive — triggers email
            notifications to every eligible student.
    """
    serializer_class = DriveSerializer

    def get_permissions(self):
        if self.request.method == "POST":
            return [IsAuthenticated(), IsPlacementManager()]
        return [IsAuthenticated()]

    def get_queryset(self):
        # Everyone (student or PM) sees the same open, not-yet-expired drives.
        # Eligibility is now surfaced as a per-drive flag (is_eligible) instead
        # of hiding drives outright — so students can see a company is visiting
        # even for other branches/batches, with Apply disabled if they don't match.
        # A drive drops off this list on its own the moment its deadline passes.
        return Drive.objects.filter(
            status=Drive.Status.OPEN,
            last_date_of_application__gte=timezone.now(),
        )

    def perform_create(self, serializer):
        drive = serializer.save(posted_by=self.request.user)
        self._notify_eligible_students(drive)

    def _notify_eligible_students(self, drive):
        eligible_students = StudentProfile.objects.filter(
            course__in=drive.eligible_courses,
            batch__in=drive.eligible_batches,
        ).select_related("user")

        courses_str = "/".join(c.replace("B.Tech - ", "") for c in drive.eligible_courses)
        batches_str = "/".join(drive.eligible_batches)
        subject = f"Placement Opportunity | {drive.company_name} {drive.drive_type} | {courses_str} {batches_str} Batch | {drive.ctc}"

        for profile in eligible_students:
            if not profile.user or not profile.user.email:
                continue

            lines = [
                "Dear Student,",
                "",
                "Greetings from the Training & Placement Cell!",
                "",
                f"We are pleased to inform you about an exciting {drive.drive_type} Opportunity with {drive.company_name}"
                f" for the {batches_str} Batch.",
                "",
                "Drive Details:",
                f"Company Name: {drive.company_name}",
                f"Drive Type: {drive.drive_type}",
                f"Profile Offered: {', '.join(drive.profiles_offered)}",
                f"Salary Package: {drive.ctc}",
                f"Job Location: {drive.job_location}",
                f"Eligible Courses: {', '.join(drive.eligible_courses)}",
                f"Eligible Batch: {batches_str}",
                f"Last Date to Apply: {drive.last_date_of_application.strftime('%d %B %Y')}",
            ]

            if drive.company_website:
                lines += ["", "Company Website:", drive.company_website]
            if drive.jd_text:
                lines += ["", "Job Description: available on the portal — open this drive and click \"Open JD\"."]

            lines += ["", "Apply directly on the SAITM Placement Portal — log in and go to Jobs & Placements."]

            if drive.company_link:
                lines += [
                    "",
                    f"Important: This company also requires registration on their own portal: {drive.company_link}",
                    "Please complete both the in-portal application and the company's own registration.",
                ]

            if drive.pm_note:
                lines += ["", f"Note: {drive.pm_note}"]

            lines += ["", "All the best!", "— Placement Manager, T&P Cell, SAITM"]

            try:
                send_mail(
                    subject=subject,
                    message="\n".join(lines),
                    from_email=None,
                    recipient_list=[profile.user.email],
                    fail_silently=True,
                )
            except Exception:
                continue


class MyDrivesView(generics.ListAPIView):
    """Drives Floated — only drives posted by the logged-in PM."""
    serializer_class = DriveSerializer
    permission_classes = [IsAuthenticated, IsPlacementManager]

    def get_queryset(self):
        return Drive.objects.filter(posted_by=self.request.user)


class ApplyToDriveView(APIView):
    """
    POST /api/drives/<id>/apply/
    This is the ONLY place an Application record gets created — meaning this
    is also the only place "Applications sent" or "Applied" status changes.
    Opening the form, filling fields, even uploading files does nothing until
    this endpoint is actually called (i.e. the student clicked Submit).

    Server re-checks, regardless of what the frontend already validated:
      - drive is still open
      - student's course + batch still match the drive's eligibility
      - deadline hasn't passed
      - every mandatory Google-Form-equivalent field is present
    """
    permission_classes = [IsAuthenticated, IsStudent]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request, pk):
        try:
            drive = Drive.objects.get(pk=pk, status=Drive.Status.OPEN)
        except Drive.DoesNotExist:
            return Response({"detail": "Drive not found or closed."}, status=status.HTTP_404_NOT_FOUND)

        profile = getattr(request.user, "student_profile", None)
        if not profile:
            return Response({"detail": "Complete your student profile first."}, status=status.HTTP_400_BAD_REQUEST)

        if profile.course not in drive.eligible_courses or profile.batch not in drive.eligible_batches:
            return Response({"detail": "You are not eligible for this drive."}, status=status.HTTP_403_FORBIDDEN)

        if timezone.now() > drive.last_date_of_application:
            return Response({"detail": "Application deadline has passed."}, status=status.HTTP_400_BAD_REQUEST)

        if Application.objects.filter(student=request.user, drive=drive).exists():
            return Response({"detail": "You have already applied to this drive."}, status=status.HTTP_400_BAD_REQUEST)

        form = ApplicationFormSerializer(data=request.data)
        form.is_valid(raise_exception=True)
        data = form.validated_data

        # roll_no is unique across all students — check before saving, so a
        # typo/collision returns a clean 400 instead of crashing the request.
        if data["roll_no"] != profile.roll_no:
            if StudentProfile.objects.filter(roll_no=data["roll_no"]).exclude(pk=profile.pk).exists():
                return Response({"roll_no": ["This roll number is already registered to another student."]}, status=status.HTTP_400_BAD_REQUEST)

        # Identity fields — editable here, saved back to the profile.
        profile.full_name = data["full_name"]
        profile.roll_no = data["roll_no"]
        profile.phone = data["phone"]
        profile.college_email = data["college_email"]
        profile.course = data["course"]
        profile.batch = data["batch"]

        # Save the reusable/profile-persisting fields back to StudentProfile,
        # so the next application this student makes comes pre-filled.
        profile.gender = data["gender"]
        profile.date_of_birth = data["date_of_birth"]
        profile.tenth_percentage = data["tenth_percentage"]
        profile.tenth_board = data["tenth_board"]
        profile.tenth_year_of_passing = data["tenth_year_of_passing"]
        profile.twelfth_percentage = data["twelfth_percentage"]
        profile.twelfth_board = data["twelfth_board"]
        profile.twelfth_year_of_passing = data.get("twelfth_year_of_passing", "")
        profile.graduation_course = data.get("graduation_course", "")
        profile.graduation_percentage = data.get("graduation_percentage", "")
        profile.current_semester_percentage = data["current_semester_percentage"]
        profile.backlogs = data["backlogs"]
        profile.has_education_gap = data["has_education_gap"]
        profile.current_location = data["current_location"]
        profile.hometown_location = data["hometown_location"]
        profile.has_internship_experience = data["has_internship_experience"]
        profile.internship_months = data.get("internship_months")
        profile.save()

        # Per-application fields — always fresh, never carried over.
        application = Application.objects.create(
            student=request.user,
            drive=drive,
            campus_name=data["campus_name"],
            aadhar_no=data["aadhar_no"],
            aadhar_file=data["aadhar_file"],
            resume_file=data["resume_file"],
        )

        # If the student browsed-and-uploaded fresh (rather than picking an
        # existing saved document), the frontend flags it here so it also
        # gets added to their reusable document library for next time.
        self._maybe_save_to_library(profile, request, "resume", "resume_file", application.resume_file)
        self._maybe_save_to_library(profile, request, "aadhar", "aadhar_file", application.aadhar_file)

        return Response(ApplicationSerializer(application).data, status=status.HTTP_201_CREATED)

    def _maybe_save_to_library(self, profile, request, doc_type, source_field_name, file_field):
        should_save = request.data.get(f"save_{doc_type}_to_profile") == "true"
        if not should_save:
            return
        existing_count = StudentDocument.objects.filter(student=profile, doc_type=doc_type).count()
        limit = MAX_DOCS_BY_TYPE.get(doc_type, DEFAULT_MAX_DOCS)
        if existing_count >= limit:
            return  # silently skip — application already succeeded, this is just a convenience save
        uploaded_file = request.FILES.get(source_field_name)
        if not uploaded_file:
            return
        StudentDocument.objects.create(
            student=profile,
            doc_type=doc_type,
            file=file_field,
            original_filename=uploaded_file.name,
        )


class ApplicationListView(generics.ListAPIView):
    """Students Applied page (PM only). Filters: ?course=...&batch=...&company=..."""
    serializer_class = ApplicationSerializer
    permission_classes = [IsAuthenticated, IsPlacementManager]

    def get_queryset(self):
        qs = Application.objects.select_related("student__student_profile", "drive")
        course = self.request.query_params.get("course")
        batch = self.request.query_params.get("batch")
        company = self.request.query_params.get("company")
        roll_no = self.request.query_params.get("roll_no")
        if course:
            qs = qs.filter(student__student_profile__course=course)
        if batch:
            qs = qs.filter(student__student_profile__batch=batch)
        if company:
            qs = qs.filter(drive__company_name=company)
        if roll_no:
            qs = qs.filter(student__student_profile__roll_no__icontains=roll_no)
        return qs


class ExportApplicationsView(APIView):
    """
    GET /api/applications/export/?course=&batch=&company=
    PM only. Downloads the SAME filtered set as the Students Applied page,
    but as a real .xlsx with every field the student filled in the
    application form — not just what's shown in the table.
    """
    permission_classes = [IsAuthenticated, IsPlacementManager]

    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse

        qs = Application.objects.select_related("student__student_profile", "drive")
        course = request.query_params.get("course")
        batch = request.query_params.get("batch")
        company = request.query_params.get("company")
        if course:
            qs = qs.filter(student__student_profile__course=course)
        if batch:
            qs = qs.filter(student__student_profile__batch=batch)
        if company:
            qs = qs.filter(drive__company_name=company)

        headers = [
            "Student Name", "Roll No", "Course", "Batch", "College Email", "Phone",
            "Company", "Drive Type", "Status", "Applied On",
            "Gender", "Date of Birth",
            "10th %", "10th Board", "10th Year", "12th %", "12th Board", "12th Year",
            "Graduation Course", "Graduation %", "Current Semester %", "Backlogs",
            "Education Gap", "Current Location", "Hometown Location",
            "Internship Experience", "Internship Months",
            "Campus Name", "Aadhar No", "Aadhar File", "Resume File",
        ]

        wb = Workbook()
        ws = wb.active
        ws.title = "Applications"
        ws.append(headers)
        for cell in ws[1]:
            cell.font = cell.font.copy(bold=True)

        for app in qs:
            profile = getattr(app.student, "student_profile", None)
            aadhar_url = request.build_absolute_uri(app.aadhar_file.url) if app.aadhar_file else ""
            resume_url = request.build_absolute_uri(app.resume_file.url) if app.resume_file else ""

            ws.append([
                profile.full_name if profile else "",
                profile.roll_no if profile else "",
                profile.course if profile else "",
                profile.batch if profile else "",
                profile.college_email if profile else "",
                profile.phone if profile else "",
                app.drive.company_name,
                app.drive.drive_type,
                app.status,
                app.applied_on.strftime("%d-%b-%Y %I:%M %p") if app.applied_on else "",
                profile.gender if profile else "",
                profile.date_of_birth.strftime("%d-%b-%Y") if profile and profile.date_of_birth else "",
                profile.tenth_percentage if profile else "",
                profile.tenth_board if profile else "",
                profile.tenth_year_of_passing if profile else "",
                profile.twelfth_percentage if profile else "",
                profile.twelfth_board if profile else "",
                profile.twelfth_year_of_passing if profile else "",
                profile.graduation_course if profile else "",
                profile.graduation_percentage if profile else "",
                profile.current_semester_percentage if profile else "",
                profile.backlogs if profile else "",
                ("Yes" if profile.has_education_gap else "No") if profile and profile.has_education_gap is not None else "",
                profile.current_location if profile else "",
                profile.hometown_location if profile else "",
                ("Yes" if profile.has_internship_experience else "No") if profile and profile.has_internship_experience is not None else "",
                profile.internship_months if profile and profile.internship_months else "",
                app.campus_name,
                app.aadhar_no,
                aadhar_url,
                resume_url,
            ])

        for i, header in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(i)].width = max(14, min(len(header) + 4, 32))

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="applications_export.xlsx"'
        wb.save(response)
        return response


class MyApplicationsView(generics.ListAPIView):
    """A student's own applications — used by the Jobs & Placements page to know which drives are already Applied."""
    serializer_class = ApplicationSerializer
    permission_classes = [IsAuthenticated, IsStudent]

    def get_queryset(self):
        return Application.objects.filter(student=self.request.user)


class ResumeSampleTemplateListCreateView(generics.ListCreateAPIView):
    """
    GET  -> anyone authenticated (students browse formats, PM manages them).
    POST -> Placement Manager only.
    """
    serializer_class = ResumeSampleTemplateSerializer
    parser_classes = [MultiPartParser, FormParser]

    def get_permissions(self):
        if self.request.method == "POST":
            return [IsAuthenticated(), IsPlacementManager()]
        return [IsAuthenticated()]

    def get_queryset(self):
        return ResumeSampleTemplate.objects.all()

    def perform_create(self, serializer):
        uploaded_file = self.request.data.get("file")
        original_name = uploaded_file.name if uploaded_file else ""
        serializer.save(uploaded_by=self.request.user, original_filename=original_name)


class ResumeSampleTemplateDeleteView(APIView):
    """DELETE /api/resume-formats/<id>/ — PM only, own uploads only."""
    permission_classes = [IsAuthenticated, IsPlacementManager]

    def delete(self, request, pk):
        try:
            template = ResumeSampleTemplate.objects.get(pk=pk)
        except ResumeSampleTemplate.DoesNotExist:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)
        if template.uploaded_by_id != request.user.id:
            return Response({"detail": "You can only remove formats you uploaded."}, status=status.HTTP_403_FORBIDDEN)
        template.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)